from __future__ import annotations

from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.main import app, get_db_session, pwd_context
from models.db import Base
from models.schema import Role, User
from parser.man09_excel_parser import Man09ExcelParser, looks_like_man09_excel


def _write_man09_workbook(path: Path, *, company: str = "DELTA TANKERS LTD") -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Personal Data"
    ws["A2"] = company
    ws["A3"] = "FORM: MAN 09"
    ws["A7"] = "SEAFARER'S EMPLOYMENT APPLICATION FORM"
    ws["A8"] = "OIL TANKERS"
    ws["A13"] = "RANK*:"
    ws["D13"] = "2/E"
    ws["G13"] = "DATE OF APPLICATION*"
    ws["I13"] = "2026-06-01"
    ws["G14"] = "AVAILABLE FROM"
    ws["I14"] = "2026-07-15"
    ws["A15"] = "SURNAME*:"
    ws["D15"] = "Buryk"
    ws["G15"] = "DATE OF BIRTH:"
    ws["I15"] = "1988-07-31"
    ws["A16"] = "FIRST NAME*:"
    ws["D16"] = "Andrii"
    ws["G16"] = "PLACE OF BIRTH:"
    ws["I16"] = "Czech Republic"
    ws["A18"] = "FATHER'S NAME:"
    ws["D18"] = "Valerii"
    ws["G18"] = "AGE:"
    ws["I18"] = "37"
    ws["A19"] = "MOTHER'S NAME:"
    ws["D19"] = "Nadiya"
    ws["G19"] = "NATIONALITY*:"
    ws["I19"] = "UKRAINIAN"
    ws["A20"] = "DIPLOMA"
    ws["D20"] = "2ND ENGINEER"
    ws["G20"] = "DIPLOMA NO:"
    ws["I20"] = "13884/2024"
    ws["A21"] = "DIPLOMA DATE:"
    ws["D21"] = "2024-03-12"
    ws["A22"] = "MARINE ACADEMY"
    ws["D22"] = "National University Odessa Maritime Academy"
    ws["G23"] = "EXPIRATION DATE"
    ws["I23"] = "2027-01-05"
    ws["A24"] = "SEAMAN BOOK No.:"
    ws["D24"] = "AB 646692"
    ws["G24"] = "PASSPORT No.:"
    ws["I24"] = "FJ969714"
    ws["A25"] = "DATE OF ISSUE:"
    ws["D25"] = "2018-11-09"
    ws["G25"] = "DATE OF ISSUE:"
    ws["I25"] = "2017-11-14"
    ws["A26"] = "DATE OF EXPIRY:"
    ws["D26"] = "2028-11-08"
    ws["G26"] = "DATE OF EXPIRY:"
    ws["I26"] = "2027-11-14"
    ws["A27"] = "PLACE OF ISSUE:"
    ws["D27"] = "Port Odessa"
    ws["G27"] = "PLACE OF ISSUE:"
    ws["I27"] = "Odessa, Ukraine"
    ws["A28"] = "RESIDENT ADDRESS"
    ws["D28"] = "2 Mala Arnautska str"
    ws["E28"] = "Odesa"
    ws["F28"] = "65000"
    ws["G28"] = "TEMPORARY ADDRESS:"
    ws["I28"] = "Constanta"
    ws["A30"] = "PHONE No.:"
    ws["D30"] = "+380933113920"
    ws["G30"] = "NAME NEXT OF KIN:"
    ws["I30"] = "Buryk Kateryna"
    ws["A31"] = "MOBILE No.:"
    ws["D31"] = "+380933113921"
    ws["G31"] = "RELATIONSHIP:"
    ws["I31"] = "Wife"
    ws["A32"] = "EMAIL ADDRESS"
    ws["D32"] = "mailto:lestad20061@gmail.com / other@example.com"
    ws["G32"] = "PHONE:"
    ws["I32"] = "+380630272266"
    ws["A33"] = "MARITAL STATUS:"
    ws["D33"] = "Married"
    ws["A35"] = "NAME:"
    ws["D35"] = "Buryk Alina"
    ws["A42"] = "USA VISA EXP.DATE:"
    ws["D42"] = "2028-12-10"
    ws["G42"] = "Height:"
    ws["I42"] = "178"
    ws["G43"] = "Weight:"
    ws["I43"] = "100"

    service = wb.create_sheet("Previous Services-Certificates")
    service["A2"] = company
    service["A7"] = "PREVIOUS SEA SERVICE"
    headers = [
        "SHIP'S NAME*",
        "YEAR OF BUILT",
        "FLAG",
        "OWNER",
        "AGENCY",
        "COUNTRY",
        "DWT",
        "SHIP TYPE*",
        "ENGINE TYPE",
        "ECDIS/DG MAKER",
        "SIGN-ON*",
        "SIGN-OFF*",
        "RANK*",
        "PERIOD ONBOARD",
        "PERIOD ASHORE",
        "REASON OF LEAVING/ PORT",
    ]
    for idx, header in enumerate(headers, start=1):
        service.cell(row=9, column=idx, value=header)
    row = [
        "Chemtrans Baltic",
        "2005",
        "Marshall Islands",
        "Chemikalien",
        "STMA",
        "Germany",
        "73897",
        "Oil Tanker",
        "MAN B&W",
        "Yanmar",
        "2025-03-26",
        "2025-06-14",
        "2/E",
        "2M 19D",
        "0M 21D",
        "EOC",
    ]
    for idx, value in enumerate(row, start=1):
        service.cell(row=10, column=idx, value=value)
    service["A12"] = "CERTIFICATES"
    service["A13"] = "BASIC SAFETY Section A-VI/1"
    service["K13"] = "ECDIS TYPE SPECIFIC TRAINING"

    wb.create_sheet("Comments")
    wb.create_sheet("Education History-Signatures")
    wb.save(path)
    return path


def test_man09_excel_parser_extracts_personal_documents_and_service(tmp_path: Path) -> None:
    workbook = _write_man09_workbook(tmp_path / "delta_man09.xlsx")

    assert looks_like_man09_excel(workbook) is True

    result = Man09ExcelParser().parse(workbook)
    personal = result["personal_data"]

    assert personal["company_name"] == "Delta Tankers"
    assert personal["surname"] == "Buryk"
    assert personal["first_name"] == "Andrii"
    assert personal["date_of_birth"] == "1988-07-31"
    assert personal["place_of_birth"] == "Czech Republic"
    assert personal["citizenship"] == "UKRAINIAN"
    assert personal["height_cm"] == "178"
    assert personal["weight_kg"] == "100"
    assert personal["email"] == "lestad20061@gmail.com"
    assert personal["secondary_email"] == "other@example.com"

    passport = next(item for item in result["documents"] if item["document_type"] == "Passport")
    assert passport["document_number"] == "FJ969714"
    assert passport["date_of_expiry"] == "2027-11-14"

    assert result["family_contacts"][0]["full_name"] == "Buryk Kateryna"
    assert result["applications"][0]["rank_applied_for"] == "2/E"

    sea = result["sea_service"][0]
    assert sea["vessel_name"] == "Chemtrans Baltic"
    assert sea["year_built"] == 2005
    assert sea["dwt"] == "73897"
    assert sea["sign_on_date"] == "2025-03-26"

    coc = result["certificates"][0]
    assert coc["certificate_number"] == "13884/2024"
    assert coc["expiry_date"] == "2027-01-05"


def test_upload_man09_excel_assigns_company(tmp_path: Path) -> None:
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    Base.metadata.create_all(bind=engine)

    def override_get_db_session():
        session = TestingSessionLocal()
        try:
            yield session
        finally:
            session.close()

    app.dependency_overrides[get_db_session] = override_get_db_session

    try:
        with TestingSessionLocal() as session:
            admin_role = Role(name="admin", description="Full access")
            session.add(admin_role)
            session.flush()
            session.add(
                User(
                    username="admin_man09",
                    password_hash=pwd_context.hash("admin123"),
                    full_name="Test Admin",
                    role_id=admin_role.role_id,
                    is_active=True,
                )
            )
            session.commit()

        client = TestClient(app)
        login_response = client.post("/auth/login", json={"username": "admin_man09", "password": "admin123"})
        assert login_response.status_code == 200
        headers = {"Authorization": f"Bearer {login_response.json()['access_token']}"}

        companies_response = client.get("/companies-manager", headers=headers)
        assert companies_response.status_code == 200
        delta = next(item for item in companies_response.json()["companies"] if item["name"] == "Delta Tankers")

        workbook = _write_man09_workbook(tmp_path / "2E Buryk CV.xlsx")
        with workbook.open("rb") as fh:
            upload_response = client.post(
                "/upload",
                files={
                    "file": (
                        "2E Buryk CV.xlsx",
                        fh,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                headers=headers,
            )
        assert upload_response.status_code == 200, upload_response.text
        upload_payload = upload_response.json()
        candidate_id = upload_payload["candidate_id"]
        assert upload_payload["result"]["personal_data"]["source_file_name"] == "2E Buryk CV.xlsx"

        candidate_response = client.get(f"/candidates/{candidate_id}", headers=headers)
        assert candidate_response.status_code == 200
        candidate = candidate_response.json()["candidate"]
        assert candidate["company_id"] == delta["company_id"]
    finally:
        app.dependency_overrides.clear()
