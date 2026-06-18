from __future__ import annotations

import io
import zipfile

from docx import Document
from openpyxl import Workbook, load_workbook
from PIL import Image

from app.template_renderer import render_template_to_file


def _write_photo(path) -> None:
    output = io.BytesIO()
    Image.new("RGB", (240, 320), "navy").save(output, format="JPEG")
    path.write_bytes(output.getvalue())


def test_candidate_photo_placeholder_renders_into_docx(tmp_path):
    photo_path = tmp_path / "portrait.jpg"
    _write_photo(photo_path)
    template_path = tmp_path / "photo_template.docx"
    output_path = tmp_path / "photo_result.docx"
    document = Document()
    document.add_paragraph("{{ candidate_photo }}")
    document.save(template_path)

    render_template_to_file(template_path, {"candidate_photo_path": str(photo_path)}, output_path)

    with zipfile.ZipFile(output_path) as archive:
        assert any(name.startswith("word/media/") for name in archive.namelist())
        document_xml = archive.read("word/document.xml")
        assert b"candidate_photo" not in document_xml


def test_candidate_photo_placeholder_renders_into_xlsx(tmp_path):
    photo_path = tmp_path / "portrait.jpg"
    _write_photo(photo_path)
    template_path = tmp_path / "photo_template.xlsx"
    output_path = tmp_path / "photo_result.xlsx"
    workbook = Workbook()
    workbook.active["B2"] = "{{ candidate_photo }}"
    workbook.save(template_path)

    render_template_to_file(template_path, {"candidate_photo_path": str(photo_path)}, output_path)

    rendered = load_workbook(output_path)
    assert rendered.active["B2"].value is None
    assert len(rendered.active._images) == 1
