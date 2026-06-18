from __future__ import annotations

from copy import deepcopy
from pathlib import Path
from typing import Any
from uuid import uuid4

from docx.shared import Mm
from docxtpl import DocxTemplate, InlineImage
from jinja2 import Environment, StrictUndefined
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

from app.attachment_naming import safe_file_part

DOCX_TEMPLATE_SUFFIXES = frozenset({".docx"})
EXCEL_TEMPLATE_SUFFIXES = frozenset({".xlsx", ".xlsm"})
SUPPORTED_RENDER_TEMPLATE_SUFFIXES = DOCX_TEMPLATE_SUFFIXES | EXCEL_TEMPLATE_SUFFIXES

TEMPLATE_MEDIA_TYPES = {
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
}

_JINJA_ENV = Environment(undefined=StrictUndefined, autoescape=False, finalize=lambda value: "" if value is None else value)


def template_media_type(template_path: Path) -> str:
    return TEMPLATE_MEDIA_TYPES.get(template_path.suffix.lower(), "application/octet-stream")


def is_renderable_template_path(template_path: Path) -> bool:
    return template_path.suffix.lower() in SUPPORTED_RENDER_TEMPLATE_SUFFIXES


def build_generated_template_name(candidate: Any, template_stem: str, suffix: str) -> str:
    applications = getattr(candidate, "applications", None) or []
    first_application = applications[0] if applications else None
    raw_position = (
        (getattr(first_application, "position_applied_for", None) if first_application else None)
        or getattr(candidate, "current_rank", None)
        or "position"
    )
    ext = suffix.lower() if suffix.startswith(".") else f".{suffix.lower()}" if suffix else ".docx"
    return (
        f"{safe_file_part(raw_position)}_"
        f"{safe_file_part(getattr(candidate, 'surname', None) or 'surname')}_"
        f"{safe_file_part(getattr(candidate, 'first_name', None) or 'name')}_"
        f"{safe_file_part(template_stem)}_"
        f"{uuid4().hex[:8]}{ext}"
    )


def render_docx_template(template_path: Path, context: dict[str, Any], output_path: Path) -> None:
    from app.docx_template_jinja import strip_email_hyperlinks_from_docx
    from app.template_field_values import prepare_docx_template_context

    render_context = prepare_docx_template_context(context, template_path)
    doc = DocxTemplate(str(template_path))
    raw_photo_path = render_context.pop("candidate_photo_path", "")
    photo_path = Path(str(raw_photo_path)) if raw_photo_path else None
    photo_value: Any = ""
    if photo_path and photo_path.is_file():
        photo_value = InlineImage(doc, str(photo_path), width=Mm(35))
    render_context["candidate_photo"] = photo_value
    render_context["photo"] = photo_value
    doc.render(render_context)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc.save(str(output_path))
    strip_email_hyperlinks_from_docx(output_path)


def _render_excel_text(value: str, context: dict[str, Any]) -> str:
    if "{{" not in value and "{%" not in value and "{#" not in value:
        return value
    rendered = _JINJA_ENV.from_string(value).render(context)
    if not value.startswith("=") and rendered.startswith("="):
        return "'" + rendered
    return rendered


def render_excel_template(template_path: Path, context: dict[str, Any], output_path: Path) -> None:
    from app.template_field_values import sanitize_email_values_for_template_render, sanitize_records_for_template_render

    render_context = deepcopy(context)
    raw_photo_path = render_context.pop("candidate_photo_path", "")
    photo_path = Path(str(raw_photo_path)) if raw_photo_path else None
    sanitize_email_values_for_template_render(render_context)
    sanitize_records_for_template_render(render_context)
    keep_vba = template_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(str(template_path), data_only=False, keep_vba=keep_vba)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if cell.value.strip() in {"{{ candidate_photo }}", "{{candidate_photo}}", "{{ photo }}", "{{photo}}"}:
                        cell.value = ""
                        if photo_path and photo_path.is_file():
                            image = ExcelImage(str(photo_path))
                            scale = min(132 / image.width, 170 / image.height, 1)
                            image.width = int(image.width * scale)
                            image.height = int(image.height * scale)
                            worksheet.add_image(image, cell.coordinate)
                        continue
                    cell.value = _render_excel_text(cell.value, render_context)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output_path))


def render_template_to_file(template_path: Path, context: dict[str, Any], output_path: Path) -> None:
    suffix = template_path.suffix.lower()
    if suffix in DOCX_TEMPLATE_SUFFIXES:
        render_docx_template(template_path, context, output_path)
        return
    if suffix in EXCEL_TEMPLATE_SUFFIXES:
        render_excel_template(template_path, context, output_path)
        return
    supported = ", ".join(sorted(SUPPORTED_RENDER_TEMPLATE_SUFFIXES))
    raise ValueError(f"Unsupported template type {suffix or 'unknown'}; supported: {supported}")
