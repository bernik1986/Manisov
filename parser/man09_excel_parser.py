from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from parser.base import BaseParser


_EMPTY_TOKENS = {"", "-", "--", "n/a", "na", "none", "#ref!", "#value!", "#n/a"}
_EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)


class Man09ExcelParser(BaseParser):
    """Parser for Delta Tankers / Marmaras MAN 09 Excel application forms."""

    PERSONAL_SHEET = "Personal Data"
    SERVICE_SHEET = "Previous Services-Certificates"
    FORM_TYPE = "MAN09 Excel"

    COMPANY_ALIASES = {
        "delta tankers ltd": "Delta Tankers",
        "delta tankers": "Delta Tankers",
        "marmaras navigation ltd": "Marmaras",
        "marmaras navigation": "Marmaras",
        "marmaras": "Marmaras",
    }

    @classmethod
    def looks_like(cls, file_path: str | Path) -> bool:
        path = Path(file_path)
        if path.suffix.lower() != ".xlsx":
            return False
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception:
            return False
        try:
            if cls.PERSONAL_SHEET not in wb.sheetnames or cls.SERVICE_SHEET not in wb.sheetnames:
                return False
            ws = wb[cls.PERSONAL_SHEET]
            sample: list[str] = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 15), values_only=True):
                for value in row:
                    text = _clean_text(value)
                    if text:
                        sample.append(text.lower())
            blob = " ".join(sample)
            return "form: man 09" in blob and "seafarer" in blob
        finally:
            wb.close()

    def parse(self, file_path: str | Path) -> dict[str, Any]:
        path = Path(file_path)
        extension = self.detect_format(path)
        if extension != ".xlsx":
            raise ValueError("Man09ExcelParser supports only .xlsx files")

        wb = load_workbook(path, data_only=True)
        try:
            result = self.empty_result()
            personal_ws = wb[self.PERSONAL_SHEET]
            service_ws = wb[self.SERVICE_SHEET]

            personal, documents, family_contacts = self._parse_personal_sheet(personal_ws)
            service_rows, certificate_rows = self._parse_service_sheet(service_ws, personal)

            result["personal_data"].update(personal)
            result["documents"].extend(documents)
            result["certificates"].extend(certificate_rows)
            result["sea_service"].extend(service_rows)
            result["family_contacts"].extend(family_contacts)

            rank = personal.get("current_rank") or personal.get("certificate_of_competency_rank")
            application: dict[str, Any] = {}
            if rank:
                application["position_applied_for"] = rank
                application["rank_applied_for"] = rank
            if personal.get("date_applied"):
                application["date_applied"] = personal["date_applied"]
            if personal.get("date_available"):
                application["date_available"] = personal["date_available"]
            if application:
                result["applications"].append(application)

            return self.ensure_result_contract(result)
        finally:
            wb.close()

    def _parse_personal_sheet(
        self,
        ws: Worksheet,
    ) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
        personal: dict[str, Any] = {
            "source_form_type": self.FORM_TYPE,
            "record_status": "active",
        }
        documents: list[dict[str, Any]] = []
        family_contacts: list[dict[str, Any]] = []

        raw_company = _cell(ws, "A2")
        company_name = self._company_name(raw_company)
        if company_name:
            personal["company_name"] = company_name
        if raw_company:
            personal["source_company_name"] = raw_company

        raw_fleet = _cell(ws, "A8")
        if raw_fleet:
            self._apply_fleet_summary(personal, raw_fleet)

        self._put(personal, "current_rank", _cell(ws, "D13"))
        self._put(personal, "date_applied", _date_cell(ws, "I13"))
        self._put(personal, "date_available", _date_cell(ws, "I14") or _cell(ws, "I14"))
        self._put(personal, "surname", _cell(ws, "D15"))
        self._put(personal, "first_name", _cell(ws, "D16"))
        self._put(personal, "date_of_birth", _date_cell(ws, "I15"))
        self._put(personal, "place_of_birth", _cell(ws, "I16"))
        self._put(personal, "country_of_birth", _cell(ws, "I16"))
        self._put(personal, "father_name", _cell(ws, "D18"))
        self._put(personal, "age", _cell(ws, "I18"))
        self._put(personal, "mother_name", _cell(ws, "D19"))
        nationality = _cell(ws, "I19")
        self._put(personal, "nationality", nationality)
        self._put(personal, "citizenship", nationality)

        self._put(personal, "certificate_of_competency_rank", _cell(ws, "D20"))
        self._put(personal, "certificate_of_competency_number", _cell(ws, "I20"))
        self._put(personal, "school_name", _cell(ws, "D22"))
        self._put(personal, "highest_educational_attainment", _cell(ws, "D22"))

        self._put(personal, "seaman_book_number", _cell(ws, "D24"))
        self._put(personal, "passport_number", _cell(ws, "I24"))
        self._put(personal, "passport_issue_date", _date_cell(ws, "I25"))
        self._put(personal, "passport_expiry_date", _date_cell(ws, "I26"))
        self._put(personal, "passport_place_of_issue", _cell(ws, "I27"))

        address = _join_values(_cell(ws, "D28"), _cell(ws, "E28"), _cell(ws, "F28"))
        self._put(personal, "permanent_address", address)
        self._put(personal, "home_address", address)
        self._put(personal, "city", _cell(ws, "E28"))
        self._put(personal, "postal_code", _cell(ws, "F28"))
        self._put(personal, "current_address", _cell(ws, "I28"))

        self._put(personal, "primary_phone", _cell(ws, "D30"))
        self._put(personal, "mobile_phone", _cell(ws, "D31") or _cell(ws, "D30"))
        email, secondary_email = _split_emails(_cell(ws, "D32"))
        self._put(personal, "email", email)
        self._put(personal, "secondary_email", secondary_email)
        self._put(personal, "marital_status", _cell(ws, "D33"))

        child_count = self._count_children(ws)
        if child_count is not None:
            personal["number_of_children"] = child_count

        self._put(personal, "usa_visa_expiry_date", _date_cell(ws, "D42"))
        self._put(personal, "height_cm", _cell(ws, "I42"))
        self._put(personal, "weight_kg", _cell(ws, "I43"))
        marks = _join_labeled_values(
            ("Hair", _cell(ws, "I40")),
            ("Eyes", _cell(ws, "I41")),
            ("Glasses", _cell(ws, "I44")),
        )
        self._put(personal, "distinctive_marks", marks)

        full_name = _join_values(personal.get("surname"), personal.get("first_name"))
        self._put(personal, "full_name", full_name)
        self._put(personal, "latin_full_name", full_name)

        self._add_document(
            documents,
            "Seaman's Book",
            number=_cell(ws, "D24"),
            issue=_date_cell(ws, "D25"),
            expiry=_date_cell(ws, "D26"),
            place=_cell(ws, "D27"),
        )
        self._add_document(
            documents,
            "Passport",
            number=_cell(ws, "I24"),
            issue=_date_cell(ws, "I25"),
            expiry=_date_cell(ws, "I26"),
            place=_cell(ws, "I27"),
        )
        self._add_document(
            documents,
            "Identity Card",
            number=_cell(ws, "D39"),
            issue=_date_cell(ws, "D40") or _cell(ws, "D40"),
            place=_cell(ws, "D41"),
        )
        self._add_document(
            documents,
            "USA Visa",
            expiry=_date_cell(ws, "D42"),
            document_category="Visa",
        )

        kin_name = _cell(ws, "I30")
        kin_relationship = _cell(ws, "I31")
        kin_phone = _cell(ws, "I32")
        if kin_name:
            family_contacts.append(
                {
                    "contact_type": "next_of_kin",
                    "full_name": kin_name,
                    "relationship_to_candidate": kin_relationship,
                    "phone": kin_phone,
                    "is_emergency_contact": True,
                }
            )
            self._put(personal, "next_of_kin_full_name", kin_name)
            self._put(personal, "next_of_kin_relationship", kin_relationship)
            self._put(personal, "next_of_kin_phone", kin_phone)

        return personal, documents, family_contacts

    def _parse_service_sheet(
        self,
        ws: Worksheet,
        personal: dict[str, Any],
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        header_row = self._find_row_containing(ws, "SHIP", "NAME")
        certificate_row = self._find_row_containing(ws, "CERTIFICATES")
        sea_service: list[dict[str, Any]] = []

        if header_row:
            stop_row = certificate_row if certificate_row and certificate_row > header_row else ws.max_row + 1
            headers = {
                col_idx: self._service_header_to_field(_cell_at(ws, header_row, col_idx))
                for col_idx in range(1, ws.max_column + 1)
            }
            for row_idx in range(header_row + 1, stop_row):
                row = self._parse_sea_service_row(ws, row_idx, headers)
                if row:
                    sea_service.append(row)

        certificates = self._parse_certificates(ws, certificate_row, personal)
        return sea_service, certificates

    def _parse_sea_service_row(
        self,
        ws: Worksheet,
        row_idx: int,
        headers: dict[int, str | None],
    ) -> dict[str, Any] | None:
        row: dict[str, Any] = {}
        for col_idx, field in headers.items():
            if not field:
                continue
            value = _cell_at(ws, row_idx, col_idx)
            if not value:
                continue
            row[field] = value

        if not any(row.get(key) for key in ("vessel_name", "rank_on_vessel", "sign_on_date", "sign_off_date")):
            return None

        if row.get("sign_on_date"):
            row["sign_on_date"] = _to_iso_date(row["sign_on_date"]) or row["sign_on_date"]
        if row.get("sign_off_date"):
            row["sign_off_date"] = _to_iso_date(row["sign_off_date"]) or row["sign_off_date"]
        if row.get("year_built"):
            row["year_built"] = _digits(row["year_built"])
        if row.get("dwt"):
            row["dwt"] = _number_text(row["dwt"])

        ecdis = _cell_at(ws, row_idx, 10)
        reason = row.get("remarks")
        if ecdis:
            row["remarks"] = _join_labeled_values(("ECDIS/DG", ecdis), ("Reason", reason))

        return row

    def _parse_certificates(
        self,
        ws: Worksheet,
        certificate_row: int | None,
        personal: dict[str, Any],
    ) -> list[dict[str, Any]]:
        certificates: list[dict[str, Any]] = []

        coc_rank = personal.get("certificate_of_competency_rank")
        coc_no = personal.get("certificate_of_competency_number")
        coc_issue = _date_cell(ws.parent[self.PERSONAL_SHEET], "D21")
        coc_expiry = (
            _date_cell(ws.parent[self.PERSONAL_SHEET], "I23")
            or _date_cell(ws.parent[self.PERSONAL_SHEET], "I21")
        )
        if coc_rank or coc_no:
            certificates.append(
                {
                    "certificate_group": "Competency",
                    "certificate_type": "Certificate of Competence",
                    "certificate_name_raw": coc_rank or "Certificate of Competence",
                    "certificate_number": coc_no,
                    "competency_rank": coc_rank,
                    "date_issued": coc_issue,
                    "expiry_date": coc_expiry,
                    "is_present": True,
                }
            )

        if not certificate_row:
            return certificates

        seen = {self.normalize_field_label(str(item.get("certificate_type") or "")) for item in certificates}
        for row_idx in range(certificate_row + 1, ws.max_row + 1):
            for col_idx in (1, 11):
                name = _cell_at(ws, row_idx, col_idx)
                if not name:
                    continue
                normalized = self.normalize_field_label(name)
                if normalized in {"other", "certificate", "certificates"} or normalized in seen:
                    continue
                seen.add(normalized)
                certificates.append(
                    {
                        "certificate_group": "Training",
                        "certificate_type": name[:100],
                        "certificate_name_raw": name,
                        "is_present": True,
                    }
                )
        return certificates

    @staticmethod
    def _put(target: dict[str, Any], key: str, value: Any) -> None:
        if _has_value(value):
            target[key] = value

    @classmethod
    def _company_name(cls, value: Any) -> str | None:
        normalized = re.sub(r"\s+", " ", str(value or "").strip().lower())
        return cls.COMPANY_ALIASES.get(normalized)

    @staticmethod
    def _apply_fleet_summary(personal: dict[str, Any], raw_fleet: str) -> None:
        normalized = raw_fleet.strip().lower()
        if "oil" in normalized or "tanker" in normalized:
            personal["oil_tanker_experience"] = True
        if "bulk" in normalized:
            personal["bulk_carrier_experience"] = True

    @staticmethod
    def _count_children(ws: Worksheet) -> int | None:
        explicit = _cell(ws, "D34")
        if explicit is not None:
            parsed = _digits(explicit)
            if parsed is not None:
                return parsed
        names = [_cell(ws, f"D{row_idx}") for row_idx in range(35, 39)]
        count = sum(1 for name in names if name)
        return count if count else None

    @staticmethod
    def _add_document(
        documents: list[dict[str, Any]],
        document_type: str,
        *,
        number: Any = None,
        issue: Any = None,
        expiry: Any = None,
        place: Any = None,
        document_category: str = "Identity",
    ) -> None:
        if not any(_has_value(value) for value in (number, issue, expiry, place)):
            return
        documents.append(
            {
                "document_category": document_category,
                "document_type": document_type,
                "document_name_raw": document_type,
                "document_number": number,
                "date_of_issue": issue,
                "date_of_expiry": expiry,
                "place_of_issue": place,
            }
        )

    @staticmethod
    def _find_row_containing(ws: Worksheet, *tokens: str) -> int | None:
        needles = [token.lower() for token in tokens if token]
        for row in ws.iter_rows():
            row_text = " ".join(_clean_text(cell.value).lower() for cell in row if _clean_text(cell.value))
            if row_text and all(token in row_text for token in needles):
                return row[0].row
        return None

    def _service_header_to_field(self, header: str | None) -> str | None:
        normalized = self.normalize_field_label(header or "")
        if not normalized:
            return None
        if "ship" in normalized and "name" in normalized:
            return "vessel_name"
        if "year" in normalized and ("built" in normalized or "build" in normalized):
            return "year_built"
        if normalized == "flag":
            return "flag"
        if normalized == "owner":
            return "employer"
        if normalized == "agency":
            return "manning_agency"
        if normalized == "country":
            return "trade_area"
        if normalized == "dwt":
            return "dwt"
        if "ship" in normalized and "type" in normalized:
            return "vessel_type"
        if "engine" in normalized and "type" in normalized:
            return "main_engine"
        if "sign" in normalized and "on" in normalized:
            return "sign_on_date"
        if "sign" in normalized and "off" in normalized:
            return "sign_off_date"
        if normalized == "rank":
            return "rank_on_vessel"
        if "period onboard" in normalized:
            return "contract_duration"
        if "reason" in normalized or "leaving" in normalized or "port" in normalized:
            return "remarks"
        return None


def looks_like_man09_excel(file_path: str | Path) -> bool:
    return Man09ExcelParser.looks_like(file_path)


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    if text.lower() in _EMPTY_TOKENS:
        return None
    if text.startswith("${") and text.endswith("}"):
        return None
    return text


def _cell(ws: Worksheet, coordinate: str) -> str | None:
    return _clean_text(ws[coordinate].value)


def _cell_at(ws: Worksheet, row: int, column: int) -> str | None:
    return _clean_text(ws.cell(row=row, column=column).value)


def _date_cell(ws: Worksheet, coordinate: str) -> str | None:
    return _to_iso_date(ws[coordinate].value)


def _to_iso_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _clean_text(value)
    if not text:
        return None
    normalized = text.replace("\\", "/").replace("-", "/").replace(".", "/")
    for fmt in ("%Y/%m/%d", "%d/%m/%Y", "%d/%m%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(normalized, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _has_value(value: Any) -> bool:
    return _clean_text(value) is not None


def _join_values(*values: Any) -> str | None:
    parts = []
    for value in values:
        text = _clean_text(value)
        if text:
            parts.append(text)
    return ", ".join(parts) if parts else None


def _join_labeled_values(*items: tuple[str, Any]) -> str | None:
    parts = []
    for label, value in items:
        text = _clean_text(value)
        if text:
            parts.append(f"{label}: {text}")
    return "; ".join(parts) if parts else None


def _split_emails(value: Any) -> tuple[str | None, str | None]:
    text = _clean_text(value)
    if not text:
        return None, None
    text = re.sub(r"^mailto:\s*", "", text, flags=re.IGNORECASE)
    matches = _EMAIL_RE.findall(text)
    if matches:
        first = matches[0]
        second = next((item for item in matches[1:] if item.lower() != first.lower()), None)
        return first, second
    return text, None


def _digits(value: Any) -> int | None:
    text = _clean_text(value)
    if not text:
        return None
    match = re.search(r"\d+", text.replace(",", ""))
    if not match:
        return None
    return int(match.group(0))


def _number_text(value: Any) -> str | None:
    text = _clean_text(value)
    if not text:
        return None
    match = re.search(r"\d+(?:[.,]\d+)?", text.replace(" ", ""))
    return match.group(0).replace(",", ".") if match else text
